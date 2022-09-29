import os
import openpyxl as opxl
import random
import lcpunit
import sys

#As each type of database will have its own python library, and therefore, code, I'll create a class for each one of them in the future. For v1, only *xls and *.xlsx are being accepted. I need to see and test if *.csv files can be read with this same code, through openpyxl.py.
class Xls():

    def __init__(self, file_path):
        """
        Initializes an instance of the object Xls().

        Parameters
        ----------
        file_path : str
            Path to the *.xls or *.xlsx file.

        Attributes
        ----------
        Str: self.file_path; self.file_path_separator; self.dir; self.file_name.

        Obj: Openpyxl.py Workbook and Worksheet objects: self.working_book; self.working_sheet.

        Dict: self.error_message.

        Returns
        -------
        Instance of Xls().

        Example
        -------
        Xls("D:/spreadsheet.xlsx").

        """

        sys.stdout.reconfigure(encoding='utf-8')

        #Sets <file_path> attribute.
        self.file_path = file_path

        #Checks if the file path is on Mac's or Windows' standard, and sets the <self.file_path_separator> to properly substring <self.file_path> later.
        if self.file_path.__contains__("\\"):
            self.file_path_separator = "\\"
        elif self.file_path.__contains__("/"):
            self.file_path_separator = "/"

        #Sets its attributes.
        self.dir = self.file_path.rpartition(self.file_path_separator)[0]
        self.file_name = self.file_path.rpartition(self.file_path_separator)[2]

        #Creates the working book and working sheet objects as attributes of Xls().
        self.working_book = opxl.load_workbook(self.file_path)
        self.working_sheet = self.working_book.active

        #Creates an empty dictionaty to store potential future error messages.
        self.error_messages = dict()


    def get_column_index(self, field_name):
        """
        Finds a given column index within the *.xls or *.xlsx database.

        This method is needed to retrieve the exact position of the column of interest, in order to iteracte with it.

        Parameters
        ----------
        field_name : str
            Name of the column.

        Returns
        -------
        int
            Index of the column.

        Example
        -------
        self.get_column_index("upload_type").

        """

        #Iteractively search for the column name provided using <str.casefold()> to avoid case sensitive errors.
        for i in range(1, ((self.working_sheet.max_column) + 1)):
            if str(self.working_sheet.cell(1, i).value).casefold() == str(field_name).casefold():
                return self.working_sheet.cell(1, i).col_idx

    def get_row_index(self, field_name, info):
        """
        Finds a given row index within the *.xls or *.xlsx database.

        This method is needed to retrieve the exact row of interest in order to iteracte with it.

        Parameters
        ----------
        field_name : str
            Name of the column.

        info : str
            Specific value to be searched.

        Returns
        -------
        int
            Index of the row.
        OR

        bool
            False

        Example
        -------
        self.get_row_index("upload_type").

        """

        #A loop iteraction to access every row of the database, starting from the row #2 to avoid the header and adding one at the end because range() is not inclusive.
        for row in range (2, ((self.working_sheet.max_row) + 1)):

            #Returns the row number of a given paper that has the same provided info <info> for the provided column <field_name> - usually the column used is the "internal_id".
            if self.working_sheet.cell(row, self.get_column_index(field_name)).value == info:
                return row

        #If the program reaches here, it's because nothing was returned and there was no row with that given <info> for that given <field_name>. Thus, it adds this information to the <error_message> dictionary, and returns false.
        self.append_to_error_dict("Row doesn't Exist:", ("No row with [" + str(field_name) + "] value equals to [" + str(info) + "] was found."))
        return False


    #-----------------------------------------------------------------
    #WRITE a function to test if the max number of rows in the file is the same as the max number of rows with values in three random columns, except the ones that starts empty, to see if there isnt a problem with the file.
    #def check_corrompt_file(self):
    #-----------------------------------------------------------------

    def check_data(self, field_name):
        """
        Checks if a specific information (i.e., for the data field "upload_type") was provided to all papers in the *.xls or *.xlsx database.

        This method checks if any specific data field was provided to all papers on the database by iteractively checking all cells of a its specific column.

        Parameters
        ----------
        field_name : str
            Name of the data field (column).

        Returns
        -------
        bool
            True if all lines (papers) have a value for that specific column (data field); False if there are any missing values.

        Example
        -------
        self.check_data("upload_type").

        """

        #Sets the counter used in this iteraction to zero;
        count = 0

        #Iteraction adds to counter if finds a cell with no value;
        #print(field_name)
        for row in range(2, ((self.working_sheet.max_row) + 1)):
            if self.working_sheet.cell(row, self.get_column_index(field_name)).value == None:
                count += 1

        #If counter equals to zero, no empty cells was find in this column, return True.
        #print("printing count of errors found in")
        #print(field_name)
        #print(count)
        if count == 0:
            return True

        #If counter equals to the number of rows (minus one because of the header), adds a specific error message to indicate that all cells are empty for this specific column.
        elif count == ((self.working_sheet.max_row) - 1):
            self.append_to_error_dict((str(field_name) + " missing:"), "Missing all of them.")
            return False

        #Else, it means the counter equals to something higher than zero and lower than the number of rows (minus header). Thus, adds a specific error message ("... missing:") with the exact coordinate of the empty cell(s) for this specific column.
        else:
            self.append_to_error_dict((str(field_name) + " missing:"), str(self.working_sheet.cell(row, self.get_column_index(field_name))))
            return False

    #-----------------------------------------------------------------
    #I've to make this check for prereserve dois if the doi is not there, so we can work with hybrid databases as well.
    #-----------------------------------------------------------------

    def check_data_content(self, field_name, content):
        """
        Checks if the content of a specified data field matches the expected value for all papers.

        This is important to check required fields with controlled vocabulary (i.e., "upload_type").

        Parameters
        ----------
        field_name : str
            Name of the data field (column).

        content : str
            Expected value in each cell of that data field (column).

        Returns
        -------
        bool
            True, if the data field exists and it's correctly filled; False, if the data field doesn't exist in the spreadsheet, or if there are values different than the desired (<content>).

        Example
        -------
        self.check_data_content("upload_type", "publication")

        """

        #As it might have more than one field with problems, I can't quite simlpy call the <return> function because it halts the program and I wouldn't know all the cells with problems. Thus, I created this variable to hold the boolean value to be returned at the end. As I'm testing for problems, the default value is True.
        checker = True

        #Removes the key "Fields and Cells without the Same Value" in the <self.error_messages> dictionary, so I can fill it again with data reffering to a different data field.
        self.error_messages.pop("Fields and Cells without the Same Value:", None)

        #Checks if the data field already exists in the spreadsheet. If not, returns False and adds an error message.
        if self.get_column_index(field_name) == None:
            self.append_to_error_dict("Fields Doesn't Exist:", field_name)
            return False

        else:
            for row in range(2, ((self.working_sheet.max_row) + 1)):

                #Tests if the value found in any cell of the data field column is different to the specified value in <content>.
                if self.working_sheet.cell(row, self.get_column_index(field_name)).value != content:

                    #Sets <checker> to false, because it found a problem.
                    checker = False

                    #Adds to the <self.error_messages> dictionary which data field presented the problem, in which cell location, with the value that is in this specific cell.
                    self.append_to_error_dict("Fields and Cells without the Same Value:", (field_name + ", " + str(self.working_sheet.cell(row, self.get_column_index(field_name))) + ": " + str(self.working_sheet.cell(row, self.get_column_index(field_name)).value)))

        return checker


    def append_to_error_dict(self, error_type, message):
        """
        Adds an error message to the <self.error_messages> dictionary.

        This appends new messages to the error dictionary, providing an unique source for these messages and therefore, later debugging.

        Parameters
        ----------
        error_type : str
            The type or name of error to be appended.
        message : str
            The content of the error.

        Example
        -------
        self.append_to_error_dict("Error type", "this was the error").

        """

        #Checks if the error_type was already included in the error dictionary <self.error_messages>, otherwise, creates it as an empty dictionary key.
        try:
            self.error_messages[error_type]
        except:
            self.error_messages[error_type] = ""

        #If the error_type is empty, simply adds the new error message to the the error dictionary <self.error_messages>.
        if self.error_messages[error_type] == "":
            self.error_messages[error_type] = self.error_messages[error_type] + str(message)

        #Else, adds the new error message after a semi-colon.
        else:
            self.error_messages[error_type] = self.error_messages[error_type] + "; " + str(message)


    def check_PDFs_validity(self):
        """
        Checks the validity of informed *.pdf's files.

        This function checks if there is any informed PDFs within the *.xls/*.xlsx database, and if they exist, it checks if they're real files. It also writes the specific type of error regarding PDF's files on the error dictionary <self.error_messages>.

        Returns
        -------
        bool
            True if it's all good; False if there is any issue with the informed files or if they aren't informed.

        Example
        -------
        self.check_PDFs_validity().

        """

        #Sets the two variables that will be used to identify possible errors found and their types.
        missing_PDFs = False
        invalid_PDFs = False

        #Checks if there is any empty cells in the PDFs column <path_to_pdf>.
        for row in range(2, ((self.working_sheet.max_row)+1)):

            #If a given cell is empty.
            if self.working_sheet.cell(row, self.get_column_index("path_to_pdf")).value == None:

                #Sets True to the control variable for missing PDFs.
                missing_PDFs = True

                #Adds that specific cell to the "Missing PDFs" key into the error dictionary <self.error_messages>.
                self.append_to_error_dict("Missing PDFs", self.working_sheet.cell(row, self.get_column_index("path_to_pdf")))

            #Currently, it assumes that the PDFs are located in the same folder as the spreadsheet.
            #Adds the file name, or path, to the spreadsheet directory in order to get the full PDF's path and test if it's a valid file.
            #It uses Mac's standard slash, which allegedly works on Windows as well.
            elif os.path.isfile((str(self.dir) + "\\" + str(self.working_sheet.cell(row, self.get_column_index("path_to_pdf")).value))) == False:

                #Sets True to the control variable for invalid PDFs.
                invalid_PDFs = True

                #Adds that specific cell to the "Invalid PDFs" key into the error dictionary <self.error_messages>.
                print(self.working_sheet.cell(row, self.get_column_index("path_to_pdf")))
                print((str(self.dir) + "\\" + str(self.working_sheet.cell(row, self.get_column_index("path_to_pdf")).value)))
                print('---')
                self.append_to_error_dict("Invalid PDFs", self.working_sheet.cell(row, self.get_column_index("path_to_pdf")))

        #Checks if any error exists, and if so, returns False.
        if missing_PDFs == True or invalid_PDFs == True:
            return False

        #Returns True as no error was found. True means 'ok'.
        else:
            return True


    def write_new_column(self, new_field_name):
        """
        Write a new column to the spreadsheet.

        Parameters
        ----------
        new_field_name : str
            New data field to be included into the spreadsheet.

        Example
        -------
        self.write_new_column("zenodo_id")

        """

        #Sets the value of a new column equals to the provided new field name <new_field_name>.
        self.working_sheet.cell(1, ((self.working_sheet.max_column) + 1)).value = str(new_field_name)

        #Saves the file.
        self.working_book.save(self.file_path)


    def write_batch_info(self, field_name, value):
        """
        Batch write a specific information <value> into a specific column <field_name>.

        It's used on data fields that are required and/or has controlled vocabulary.

        Parameters
        ----------
        field_name : str
            Data field to batch-write the information into the spreadsheet.

        value : str
            Information that will be written multiple times in the specified data field.

        Example
        -------
        self.write_batch_info("upload_type", "publication").

        """

        #Deals differently if the data field <field_name> is equals to "internal_id", as this is used just as coordinate to write back information in the future. The argument <value> is ignored here, and a sequential number starting from 1 is added to every and each paper in the spreadsheet.
        if field_name == "internal_id":

            for row in range(2, ((self.working_sheet.max_row) + 1)):

                #Simply adds an integer (number of row minus 1) as the internal_id for all papers within a spreadsheet.
                self.working_sheet.cell(row, self.get_column_index(field_name)).value = str((row - 1))

        else:

            for row in range(2, ((self.working_sheet.max_row) + 1)):

                #If the data field <field_name> isn't equals to "internal_id", adds as cell value for every paper (row) the same specified value <value>.
                self.working_sheet.cell(row, self.get_column_index(field_name)).value = str(value)

        #Saves the file.
        self.working_book.save(self.file_path)


    def write_checking_internal_id(self, internal_id, field_name, value):
        """
        Writes a specific info <value> into a specific data field <field_name>, for a specific paper <internal_id>.

        Parameters
        ----------
        internal_id : str
            Internal ID number <internal_id> to identify which paper should receive the information <value> in the data field specified by <field_name>.

        field_name : str
            Data field <field_name> that will have the information <value> written for a specific paper <internal_id>.

        value : str
            Information that will be written multiple times in the specified data field.

        Example
        -------
        self.write_checking_internal_id("2", "zenodo_id", "26222").

        """

        #Goes through all rows of the spreadsheet.
        for row in range (2, ((self.working_sheet.max_row) + 1)):

            #Identifies which row (paper) has the same internal ID as the one specified by <internal_id>.
            if self.working_sheet.cell(row, self.get_column_index("internal_id")).value == internal_id:

                #Writes the information <value> in that specific row (paper), in a specific data field <field_name>.
                self.working_sheet.cell(row, self.get_column_index(field_name)).value = str(value)

                #Saves the file.
                self.working_book.save(self.file_path)

                #Halts the program since the internal id <internal_id> is unique and was already found.
                break


    def get_data(self, field_name, row_number):

        if field_name == "keywords" or field_name == "communities" or field_name == "creators" or field_name == "contributors":

            list = []

            if field_name == "keywords" or field_name == "communities":

                if field_name == "keywords":
                    field_name = field_name[:(len(field_name) - 1)]

                if field_name == "communities":
                    field_name = (field_name[:(len(field_name) - 3)] + "y")

                for col in range(1, ((self.working_sheet.max_column)+1)):

                  # print(self.working_sheet.cell(row_number, col).value)
                  print(self.working_sheet.cell(1, col).value)
                  print(self.working_sheet.cell(1, col))
                  if field_name in self.working_sheet.cell(1, col).value and self.working_sheet.cell(row_number, col).value != None:
                    if field_name == "keyword":
                        list.append(self.working_sheet.cell(row_number, col).value)

                    elif field_name == "community":
                        list.append({"identifier": self.working_sheet.cell(row_number, col).value})

                return list

            else:

                checker = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "0"]

                for each_num in checker:

                    for col in range (1, ((self.working_sheet.max_column)+1)):

                        if each_num in self.working_sheet.cell(1, col).value and "name" in self.working_sheet.cell(1, col).value and field_name[:(len(field_name)-1)] in self.working_sheet.cell(1, col).value:

                            dict = {}

                            text = self.working_sheet.cell(1, col).value
                            num = text[(text.find("_")+1):][:text[(text.find("_")+1):].find("_")]

                            col_indexes = {
                                "name": self.get_column_index((field_name[:(len(field_name)-1)] + "_" + num + "_name")),
                                "affiliation": self.get_column_index((field_name[:(len(field_name)-1)] + "_" + num + "_affiliation")),
                                "orcid": self.get_column_index((field_name[:(len(field_name)-1)] + "_" + num + "_orcid")),
                                "gnd": self.get_column_index((field_name[:(len(field_name)-1)] + "_" + num + "_gnd"))
                            }

                            if field_name == "contributors":
                                col_indexes.update({"type": self.get_column_index((field_name[:(len(field_name)-1)] + "_" + num + "_type"))})


                            for each_col in col_indexes:

                                if col_indexes[each_col] != None and self.working_sheet.cell(row_number, col_indexes[each_col]).value != None:

                                    if "orcid" in each_col:
                                        dict.update({each_col: str(self.working_sheet.cell(row_number, col_indexes[each_col]).value)})

                                    else:
                                        dict.update({each_col: self.working_sheet.cell(row_number, col_indexes[each_col]).value})

                            if dict != {}:
                                list.append(dict)

                return list

        elif field_name == "notes":

            notes = ""

            fields_for_notes = ["second_language", "third_language"]

            for each_field in fields_for_notes:

                if self.get_column_index(each_field) != None and self.working_sheet.cell(row_number, self.get_column_index(each_field)).value != None:

                    language = self.working_sheet.cell(row_number, self.get_column_index(each_field)).value.casefold()

                    if self.get_column_index((language + "_title")) != None and self.working_sheet.cell(row_number, self.get_column_index((language + "_title"))).value != None:

                        notes = notes + ("[" + (self.working_sheet.cell(row_number, self.get_column_index(each_field)).value + " Title") + "]" + "\n" + (self.working_sheet.cell(row_number, self.get_column_index((language + "_title"))).value))

                    if self.get_column_index((language + "_abstract")) != None and self.working_sheet.cell(row_number, self.get_column_index((language + "_abstract"))).value != None:

                        notes = notes + "\n" + ("[" + (self.working_sheet.cell(row_number, self.get_column_index(each_field)).value + " Abstract") + "]" + "\n" + (self.working_sheet.cell(row_number, self.get_column_index((language + "_abstract"))).value))

            return notes

        elif self.get_column_index(field_name) != None:
            return self.working_sheet.cell(row_number, self.get_column_index(field_name)).value

        #elif self.get_column_index(field_name) == None:
        else:
            self.append_to_error_dict("Get Data Issue with Fields", field_name)
            return False


    def check_internal_ID(self):
        """
        Makes sure that the data field "internal_id" exists and has valid values for each paper.

        Returns
        -------
        bool
            Always returns True as this method fixes any possible problems with the data field "internal_id".

        Example
        -------
        self.check_internal_ID().

        """

        #Checks if the data field "internal_id" exists.
        if self.get_column_index("internal_id") == None:

            #If not, adds to the spreadsheet.
            self.write_new_column("internal_id")

            #Goes through all rows and adds a sequential number.
            for row in range(2, ((self.working_sheet.max_row) + 1)):

                #The number "1" as argument doesn't mean anything, because when "internal_id" is provided as argument for this method, the argument <value> is desconsidered.
                #I should CHANGE this because now I know how to use optional argumens in methods and functions. Should apply it here.
                self.write_batch_info("internal_id", 1)

        #Checks, in case of existence of the data field "internal_id", if every and each paper has a value already.
        elif self.check_data("internal_id") == False:

                #If they don't, it writes the sequential number, as explained in a comment above.
                self.write_batch_info("internal_id", 1)

        #Returns True, as it fixes all potential problems with the data field "internal_id", to indicate that it is succesfully included.
        return True


    def check_PDF_fields(self):
        """
        .

        Returns
        -------
        bool
            .

        Example
        -------
        .

        """

        list = ["path_to_pdf", "pdf_status"]

        for each in list:

            if self.get_column_index(each) == None:
                self.write_new_column(str(each))

        if self.get_column_index(list[0]) != None and self.get_column_index(list[1]) != None:
            return True

        elif self.get_column_index(list[0]) == None:
            self.append_to_error_dict("PDFs' Fields Missing", list[0])
            return False

        elif self.get_column_index(list[1]) == None:
            self.append_to_error_dict("PDFs' Fields Missing", list[1])
            return False


    def check_resource_fields(self):

        list_of_resource_fields = ["zenodo_id", "zenodo_doi", "zenodo_doi_url", "zenodo_published", "zenodo_created", "zenodo_modified", "zenodo_owner", "zenodo_record_id", "zenodo_record_url", "zenodo_status", "zenodo_prereserve_doi"]

        for each in list_of_resource_fields:
            if self.get_column_index(each) == None:
                self.write_new_column(each)

        return True


    def check_metadata_fields(self):

        issue = False

        #For articles
        list_only = ["upload_type", "publication_type", "publication_date", "title", "description", "access_right", "license", "journal_title", "journal_volume", "journal_issue", "journal_pages", "language"]

        #For conference papers
        #list_only = ["upload_type", "publication_type", "publication_date", "title", "description", "access_right", "license", "partof_title", "conference_title", "conference_dates", "conference_place", "conference_url", "conference_session", "conference_session_part", "language"]

        #For image
        #list_only = ["upload_type", "image_type", "publication_date", "title", "description", "access_right", "license"]

        # For articles and conference paper
        list_required = ["upload_type", "publication_type", "publication_date", "title", "description", "access_right", "license"]

        # For images
        #list_required = ["upload_type", "image_type", "publication_date", "title", "description", "access_right", "license"]

        dict_mult = {"creator": "False",
            "keyword": "False",
            "contributor": "False",
            "community": "False"
            }

        for each in list_only:
            if self.get_column_index(each) == None:
                self.append_to_error_dict("Missing Metadata Fields", each)

                issue = True

            if each in list_required and self.check_data(each) == False:
                print(each)
                self.append_to_error_dict("Missing Metadata Content for Required Fields", each)

                issue = True

        for each in dict_mult:

            for col in range(1, (self.working_sheet.max_column + 1)):

                if each in self.working_sheet.cell(1, col).value:
                    dict_mult[each] = True
                    break

        for row in range(2, ((self.working_sheet.max_row) + 1)):

            if self.get_data("creators", row) == []:
                self.append_to_error_dict("No creator for Paper ID:", str(self.working_sheet.cell(row, self.get_column_index("internal_ID")).value))

                issue = True

        for each in dict_mult:
            if str(dict_mult[each]) == "False":
                self.append_to_error_dict("Missing Metadata Fields", each)

                issue = True

        return not(issue)

#Gerar uma que use todas as funcoes de chacagem criada para gerar um teste final sobre os dados do xls.

    """
    def check_all_fields(self):

        if self.check_internal_ID() == True and self.check_PDF_fields() == True and self.check_resource_fields() == True and self.check_metadata_fields() == True:
            return True

        else:
            self.check_PDFs_validity()
            return False
    """
    
    def check_all_fields(self):
        return True


    def define_action(self):


        #Should think about standards, if true will always carry an error status, or if its false that will do it. I prefer the secodn, however, I've used the first as well. Should fix this for the sake of consistency.
        #Should come up with a standard dictionary error messages, and fix when to call it. Mostly, I'll be calling in this function, while appending messages in the checkers.

        #return 0

        if self.check_all_fields() == True:

            if self.check_data("zenodo_status") == False:
                if self.error_messages["zenodo_status missing:"] == "Missing all of them.":
                    #This option means that we need to push the data for the first time.
                    return 0

                else:
                    print(self.error_messages)

                    #Check data because it's likely that some of the included papers were already pushed, some weren't."

                    #In the future, we might have a checker, to see if the paper is already in there and this will become a very important feature.
                    return 1


            elif self.check_data_content("zenodo_status", "done") == True or self.check_data_content("zenodo_published", "True") == True:
                if self.check_data_content("pdf_status", "sent") == True:

                    #This is returned when all papers included were already pushed and published.
                    return 2

                #I comment off this chunk of code because I don't think will be any case where we published the document without the PDFs? But if we change our mind, here is the code to do it.
                #else:
                    #print("all done, but no PDFs were pushed")

            elif self.check_data_content("zenodo_status", "unsubmitted") == True:

                if self.check_data_content("pdf_status", "sent"):
                    if self.check_PDFs_validity() == True:

                        #Means that the data were pushed, including the PDFs, for all papers in this spreadsheet and it's, therefore, ready for publishing.
                        return 3

                    else:
                        self.append_to_error_dict("PDFs Status", "Sent but current PDF information invalid")

                        print(self.error_messages)

                        #Data it's ready for publishing, meaning that all data and PDFs were sent to Zenodo already, but oddly enough, the PDFs in this spreadsheet are not validy PDFs anymore. Better to have this checked before.
                        return 4

                elif self.check_data_content("pdf_status", None):
                    if self.check_PDFs_validity() == True:

                        #Data sent, PDFs validy, ready to push PDFs to Zenodo.
                        return 5

                    else:

                        print(self.error_messages)

                        #This means that the PDFs are not validy, but the rest of the data was sent already.
                        return 6

        else:
            #If any field presented a problem after our initial check, the code will come here.
            print(self.error_messages)


#EH SO GERAR A QUE VAI LER LINHA A LINHA COLETANDO A INFORMACAO E CRIANDO INSTANCIAS DO PAPER

    def get_all_papers(self):

        list_of_papers = []

        if self.check_all_fields() == False:
            print(self.error_messages)

        else:

            for row in range (2, ((self.working_sheet.max_row) + 1)):

                dict_temp = {"internal_id": "",
                    "path_to_pdf": "",
                    "pdf_status": "",
                    "zenodo_id": "",
                    "zenodo_doi": "",
                    "zenodo_doi_url": "",
                    "zenodo_published": "",
                    "zenodo_created": "",
                    "zenodo_modified": "",
                    "zenodo_owner": "",
                    "zenodo_record_id": "",
                    "zenodo_record_url": "",
                    "zenodo_status": "",
                    "metadata": {
                        "doi": "",
                        "zenodo_prereserve": "",
                        "title": "",
                        "upload_type": "",
                        #"image_type": "",
                        "publication_type": "",
                        "publication_date": "",
                        "description": "",
                        "access_right": "",
                        "license": "",
                        #"partof_title": "",
                        "journal_title": "",
                        "journal_volume": "",
                        "journal_issue": "",
                        "journal_pages": "",
                        #"conference_title": "",
                        #"conference_dates": "",
                        #"conference_place": "",
                        #"conference_url": "",
                        #"conference_session": "",
                        #"conference_session_part": "",
                        "language": "",
                        "communities": "",
                        "keywords": "",
                        "creators": "",
                        "contributors": "",
                        #"notes": ""
                    }
                }

                for each in dict_temp:

                    if each == "metadata":

                        for metadata in dict_temp[each]:

                            dict_temp[each][metadata] = self.get_data(metadata, row)

                    elif self.get_data(each, row) == None:
                        dict_temp[each] = ""

                    else:
                        dict_temp[each] = self.get_data(each, row)

                list_of_papers.append(lcpunit.Paper(dict_temp))


        return list_of_papers
