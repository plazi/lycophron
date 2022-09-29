import os
import requests
import json
import openpyxl as opxl
import random


def data(file_path):
    #Creates the working book and working sheet objects as attributes of Xls().
    working_book = opxl.load_workbook(file_path)

    return working_book


def column_index(file_path, field_name):

    sheet = data(file_path).active

    #Iteractively search for the column name provided using <str.casefold()> to avoid case sensitive errors.
    for i in range(1, ((sheet.max_column) + 1)):
        if str(sheet.cell(1, i).value).casefold() == str(field_name).casefold():

            return sheet.cell(1, i).col_idx


def add_field(file_path, field_name):

    sheet = data(file_path)
    sheet.active.cell(1, (sheet.active.max_column + 1)).value = field_name
    sheet.save(file_path)


def write_value(file_path, field_name, row_number, value):

    sheet = data(file_path)
    sheet.active.cell(row_number, column_index(file_path, field_name)).value = value
    sheet.save(file_path)


def get_entry(file_path, row_number):

    sheet = data(file_path).active

    entry = {}

    columns = 1
    #print(sheet.max_column)
    while (columns <= sheet.max_column):
        #print(sheet.cell(1, columns).value)
        if (sheet.cell(1, columns).value.find("<") == -1):
            if ("creator" in sheet.cell(1,columns).value):
                entry["creators"] = [{ "name": str(sheet.cell(row_number, columns).value) }]
            elif ("contributor" in sheet.cell(1, columns).value):
                entry["contributors"] = [{ "name": str(sheet.cell(row_number, columns).value), "type": "DataCurator" }]
            elif ("keyword" in sheet.cell(1, columns).value):
                l = []
                l.append(str(sheet.cell(row_number, columns).value))
                entry["keywords"] = l
            elif ("community" in sheet.cell(1, columns).value):
                entry["communities"] = [{ "identifier": str(sheet.cell(row_number, columns).value) }]
            else:
                entry[sheet.cell(1, columns).value] = str(sheet.cell(row_number, columns).value)

        columns += 1

    temp = {"metadata": entry}

    return temp


def switches(sandbox="off", marcus_mode="off"):

    switches = {}

    if sandbox == "on":
        switches["url"] = "https://sandbox.zenodo.org/api/deposit/depositions"

        if marcus_mode == "on":
            switches["key"] = "3QbSCaJML0Ein4RgGB3ruAhu5xG7sBhi6rdK5DkmcN390HMaSeY2gckrabZJ"
        else:
            switches["key"] = "6R6LCyqt7qSskQZMmN9RtnOd7KR59IH977deYBfbzJjOaXnN4cFktWHCmwfn"

    else:
        switches["url"] = "https://zenodo.org/api/deposit/depositions"

        if marcus_mode == "on":
            switches["key"] = "yoCuTRygTHb0YSQ3UQY45Rj9JfcJhWhVXXvV3ZTRAQbTuiZY7KHD2fWg8S1j"
        else:
            switches["key"] = "5izwuFcjcN9Vk7En8JNeP6PqVe2oM1KIRxVEfvCGoJ4rW6cYyck5LBIWmB1v"

    return switches


def push_data(entry):

    headers = {"Content-Type": "application/json"}

    switch = switches(sandbox="off", marcus_mode="off")

    #print(switch["url"])
    #print(switch["key"])

    zenodo = requests.post(switch["url"],
        params={'access_token': switch["key"]},
        data=json.dumps(entry),
        headers=headers)

    #print(zenodo.status_code)
    #print(zenodo.json())

    print("Zenodo Status Code: " + str(zenodo.status_code))

    if zenodo.status_code == 400:
        print(zenodo.json())
    elif zenodo.status_code == 500:
        print(zenodo.json())


    return zenodo.json()


def process_response(file_path, row_number, zenodo_json):

    write_value(file_path, "<zenodo_id>", row_number, zenodo_json["id"])
    write_value(file_path, "<zenodo_state>", row_number, zenodo_json["state"])
    write_value(file_path, "<prereserve_doi>", row_number, zenodo_json["metadata"]["prereserve_doi"]["doi"])

    print("Succesfully added # " + str(row_number))
    print("------------------------")


file_path = "C:\\Users\\poa\\Documents\\Fabricius\\Fabricius_Types-Initial_Dataset_NoUnkonwns-edited-9.xlsx"

# TESTING AREA
#print(data(file_path))
#print(column_index(data(file_path), "publication_date"))
#print(data(file_path).max_column)

#print(data(file_path).cell(1, 1).value)

#print(data(file_path).max_row)

#print(get_entry(data(file_path), 4))

#print(data(file_path).active.cell(1, 1).value)
'''
add_field(file_path, "prereserved_doi2")
print(column_index(data(file_path).active, "prereserved_doi"))
print(column_index(data(file_path).active, "prereserved_doi2"))
'''

#add_field(file_path, "internal_id")
#add_value(file_path, "internal_ID", 2, "2")

#print(column_index(file_path, "internal_id"))
#print(data(file_path).active.cell(2, column_index(file_path, "internal_id")).value)

#print(switches(sandbox="on", marcus_mode="on"))

counter = 2

#while (counter <= data(file_path).active.max_row):
while (counter <= data(file_path).active.max_row):
    print("Processing record # " + str(counter))
    process_response(file_path, counter, push_data(get_entry(file_path, counter)))
    counter += 1

#print(get_entry(file_path, 2))